"""
Forecast Output Generator
Generates Excel output matching the Yogabars forecast format:
- Channels sheet: SKU info + all channels x 3 months with values
- Consolidated sheet: SKU info + cluster totals x 3 months with SUMIFS formulas
"""
import json
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta


def generate_forecast_excel(data_json, output_path):
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    cycle_month = datetime.strptime(data["cycle_month"][:10], "%Y-%m-%d")
    month1 = cycle_month
    month2 = cycle_month + relativedelta(months=1)
    month3 = cycle_month + relativedelta(months=2)
    months = [month1, month2, month3]

    skus = data["skus"]
    channels = sorted(data["channels"], key=lambda c: c["display_order"])
    clusters = sorted(data["clusters"], key=lambda c: c["display_order"])
    forecast = data["forecast"]

    channel_names = [c["name"] for c in channels]
    cluster_names = [c["name"] for c in clusters]
    num_channels = len(channel_names)
    num_clusters = len(cluster_names)

    # Build lookup: (sku_master_sku, channel_name, month_str) -> quantity
    forecast_lookup = {}
    for f in forecast:
        key = (f["sku_master_sku"], f["channel_name"], f["forecast_month"][:7])
        forecast_lookup[key] = f["quantity"]

    # Channel -> cluster mapping
    ch_cluster = {c["name"]: c["cluster_name"] for c in channels}

    wb = Workbook()

    # ==========================================
    # SHEET 1: CHANNELS
    # ==========================================
    ws_ch = wb.active
    ws_ch.title = "Channels"

    sku_col_start = 4
    sku_headers = ["New Master SKU", "Active Status", "New FG Code", "Status",
                   "Master SKU", "FG Code", "Product Name", "Category", "Product Category"]

    header_row = 6
    data_start_row = 7

    m1_gt_col = 13
    m1_ch_start = 14
    m1_ch_end = m1_ch_start + num_channels - 1

    m2_gap_col = m1_ch_end + 1
    m2_gt_col = m2_gap_col + 1
    m2_ch_start = m2_gt_col + 1
    m2_ch_end = m2_ch_start + num_channels - 1

    m3_gap_col = m2_ch_end + 1
    m3_gt_col = m3_gap_col + 1
    m3_ch_start = m3_gt_col + 1
    m3_ch_end = m3_ch_start + num_channels - 1

    month_blocks = [
        {"gt_col": m1_gt_col, "ch_start": m1_ch_start, "ch_end": m1_ch_end, "gap_col": None, "month": month1},
        {"gt_col": m2_gt_col, "ch_start": m2_ch_start, "ch_end": m2_ch_end, "gap_col": m2_gap_col, "month": month2},
        {"gt_col": m3_gt_col, "ch_start": m3_ch_start, "ch_end": m3_ch_end, "gap_col": m3_gap_col, "month": month3},
    ]

    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    month_fill = PatternFill("solid", fgColor="E2EFDA")

    # Row 3: Month dates
    for mb in month_blocks:
        ws_ch.cell(row=3, column=mb["ch_start"], value=mb["month"])
        ws_ch.cell(row=3, column=mb["ch_start"]).number_format = 'YYYY-MM-DD'
        ws_ch.cell(row=3, column=mb["ch_start"]).font = Font(bold=True, size=11)
        ws_ch.cell(row=3, column=mb["ch_start"]).fill = month_fill

    # Row 4: Cluster names for each channel column
    for mb in month_blocks:
        for i, ch_name in enumerate(channel_names):
            col = mb["ch_start"] + i
            ws_ch.cell(row=4, column=col, value=ch_cluster.get(ch_name, ""))

    # Row 5: SUBTOTAL formulas
    last_data_row = data_start_row + len(skus) - 1
    for mb in month_blocks:
        gt_letter = get_column_letter(mb["gt_col"])
        ws_ch.cell(row=5, column=mb["gt_col"],
                   value=f'=SUBTOTAL(9,{gt_letter}{data_start_row}:{gt_letter}{last_data_row})')
        for i in range(num_channels):
            col = mb["ch_start"] + i
            col_letter = get_column_letter(col)
            ws_ch.cell(row=5, column=col,
                       value=f'=SUBTOTAL(9,{col_letter}{data_start_row}:{col_letter}{last_data_row})')

    # Row 6: Headers
    for i, h in enumerate(sku_headers):
        c = ws_ch.cell(row=header_row, column=sku_col_start + i, value=h)
        c.font = header_font
        c.fill = header_fill

    for mb in month_blocks:
        c = ws_ch.cell(row=header_row, column=mb["gt_col"], value="Grand Total")
        c.font = header_font
        c.fill = header_fill
        for i, ch_name in enumerate(channel_names):
            col = mb["ch_start"] + i
            c = ws_ch.cell(row=header_row, column=col, value=ch_name)
            c.font = header_font
            c.fill = header_fill

    # Data rows
    for sku_idx, sku in enumerate(skus):
        row = data_start_row + sku_idx
        new_master_sku = sku.get("new_master_sku", "")
        fg_code = str(sku.get("fg_code", "")).strip()
        product_name = sku.get("product_name", "")

        # Derive fields from actual DB data
        # new_master_sku is like "BB_AFG" (already has G)
        # Master SKU = remove trailing G: "BB_AF"
        master_sku = new_master_sku
        if master_sku.endswith("G"):
            master_sku = master_sku[:-1]

        # New FG Code = fg_code + "G" (e.g. "14244" -> "14244G")
        new_fg_code = f"{fg_code}G" if fg_code else ""
        # FG Code = the raw number from DB
        fg_code_display = int(fg_code) if fg_code.isdigit() else fg_code

        # SKU info columns
        ws_ch.cell(row=row, column=4, value=new_master_sku)     # New Master SKU
        ws_ch.cell(row=row, column=5, value="Active")           # Active Status
        ws_ch.cell(row=row, column=6, value=new_fg_code)        # New FG Code
        ws_ch.cell(row=row, column=7, value="")                 # Status
        ws_ch.cell(row=row, column=8, value=master_sku)         # Master SKU
        ws_ch.cell(row=row, column=9, value=fg_code_display)    # FG Code
        ws_ch.cell(row=row, column=10, value=product_name)
        ws_ch.cell(row=row, column=11, value=sku.get("category", ""))
        ws_ch.cell(row=row, column=12, value=sku.get("product_category", ""))

        for mi, mb in enumerate(month_blocks):
            month_str = months[mi].strftime("%Y-%m")

            ch_start_letter = get_column_letter(mb["ch_start"])
            ch_end_letter = get_column_letter(mb["ch_end"])
            ws_ch.cell(row=row, column=mb["gt_col"],
                       value=f'=SUM({ch_start_letter}{row}:{ch_end_letter}{row})')

            for i, ch_name in enumerate(channel_names):
                col = mb["ch_start"] + i
                key = (new_master_sku, ch_name, month_str)
                qty = forecast_lookup.get(key, 0)
                if qty and qty > 0:
                    ws_ch.cell(row=row, column=col, value=qty)

    # Column widths
    for col in range(4, 13):
        ws_ch.column_dimensions[get_column_letter(col)].width = 14
    ws_ch.column_dimensions[get_column_letter(10)].width = 45

    for mb in month_blocks:
        ws_ch.column_dimensions[get_column_letter(mb["gt_col"])].width = 14
        for i in range(num_channels):
            ws_ch.column_dimensions[get_column_letter(mb["ch_start"] + i)].width = 12

    # ==========================================
    # SHEET 2: CONSOLIDATED
    # ==========================================
    ws_con = wb.create_sheet("Consolidated")

    con_sku_start = 6
    con_sku_headers = ["New Master SKU", "New FG Code", "Master SKU", "FG Code",
                       "Product Name", "Category", "Product Category"]

    con_m1_gt = 13
    con_m1_cl_start = 14
    con_m1_cl_end = con_m1_cl_start + num_clusters - 1

    con_m2_gap = con_m1_cl_end + 1
    con_m2_gt = con_m2_gap + 1
    con_m2_cl_start = con_m2_gt + 1
    con_m2_cl_end = con_m2_cl_start + num_clusters - 1

    con_m3_gap = con_m2_cl_end + 1
    con_m3_gt = con_m3_gap + 1
    con_m3_cl_start = con_m3_gt + 1
    con_m3_cl_end = con_m3_cl_start + num_clusters - 1

    con_months = [
        {"gt_col": con_m1_gt, "cl_start": con_m1_cl_start, "cl_end": con_m1_cl_end, "gap_col": None, "month": month1},
        {"gt_col": con_m2_gt, "cl_start": con_m2_cl_start, "cl_end": con_m2_cl_end, "gap_col": con_m2_gap, "month": month2},
        {"gt_col": con_m3_gt, "cl_start": con_m3_cl_start, "cl_end": con_m3_cl_end, "gap_col": con_m3_gap, "month": month3},
    ]

    # Row 3: Month dates
    for cm in con_months:
        ws_con.cell(row=3, column=cm["gt_col"], value=cm["month"])
        ws_con.cell(row=3, column=cm["gt_col"]).number_format = 'YYYY-MM-DD'
        ws_con.cell(row=3, column=cm["gt_col"]).font = Font(bold=True, size=11)
        ws_con.cell(row=3, column=cm["gt_col"]).fill = month_fill

    # Row 4: SUBTOTAL formulas
    con_last_data_row = 5 + len(skus)
    for cm in con_months:
        gt_letter = get_column_letter(cm["gt_col"])
        ws_con.cell(row=4, column=cm["gt_col"],
                    value=f'=SUBTOTAL(9,{gt_letter}6:{gt_letter}{con_last_data_row})')
        for i in range(num_clusters):
            col = cm["cl_start"] + i
            col_letter = get_column_letter(col)
            ws_con.cell(row=4, column=col,
                        value=f'=SUBTOTAL(9,{col_letter}6:{col_letter}{con_last_data_row})')

    # Row 5: Headers
    for i, h in enumerate(con_sku_headers):
        c = ws_con.cell(row=5, column=con_sku_start + i, value=h)
        c.font = header_font
        c.fill = header_fill

    for cm in con_months:
        c = ws_con.cell(row=5, column=cm["gt_col"], value="Grand Total")
        c.font = header_font
        c.fill = header_fill
        for i, cl_name in enumerate(cluster_names):
            col = cm["cl_start"] + i
            c = ws_con.cell(row=5, column=col, value=cl_name)
            c.font = header_font
            c.fill = header_fill

    # Data rows with SUMIFS formulas
    for sku_idx, sku in enumerate(skus):
        con_row = 6 + sku_idx
        ch_row = data_start_row + sku_idx

        new_master_sku = sku.get("new_master_sku", "")
        fg_code = str(sku.get("fg_code", "")).strip()
        product_name = sku.get("product_name", "")

        master_sku = new_master_sku
        if master_sku.endswith("G"):
            master_sku = master_sku[:-1]

        new_fg_code = f"{fg_code}G" if fg_code else ""
        fg_code_display = int(fg_code) if fg_code.isdigit() else fg_code

        # SKU info
        ws_con.cell(row=con_row, column=6, value=new_master_sku)
        ws_con.cell(row=con_row, column=7, value=new_fg_code)
        ws_con.cell(row=con_row, column=8, value=master_sku)
        ws_con.cell(row=con_row, column=9, value=fg_code_display)
        ws_con.cell(row=con_row, column=10, value=product_name)
        ws_con.cell(row=con_row, column=11, value=sku.get("category", ""))
        ws_con.cell(row=con_row, column=12, value=sku.get("product_category", ""))

        for mi, cm in enumerate(con_months):
            mb = month_blocks[mi]

            cl_start_letter = get_column_letter(cm["cl_start"])
            cl_end_letter = get_column_letter(cm["cl_end"])
            ws_con.cell(row=con_row, column=cm["gt_col"],
                        value=f'=SUM({cl_start_letter}{con_row}:{cl_end_letter}{con_row})')

            ch_start_letter = get_column_letter(mb["ch_start"])
            ch_end_letter = get_column_letter(mb["ch_end"])

            for i, cl_name in enumerate(cluster_names):
                col = cm["cl_start"] + i
                cluster_col_letter = get_column_letter(col)
                formula = (
                    f"=SUMIFS(Channels!${ch_start_letter}{ch_row}:${ch_end_letter}{ch_row},"
                    f"Channels!${ch_start_letter}$4:${ch_end_letter}$4,"
                    f"Consolidated!{cluster_col_letter}$5)"
                )
                ws_con.cell(row=con_row, column=col, value=formula)

    # Column widths
    for col in range(6, 13):
        ws_con.column_dimensions[get_column_letter(col)].width = 14
    ws_con.column_dimensions[get_column_letter(10)].width = 45

    for cm in con_months:
        ws_con.column_dimensions[get_column_letter(cm["gt_col"])].width = 14
        for i in range(num_clusters):
            ws_con.column_dimensions[get_column_letter(cm["cl_start"] + i)].width = 16

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'r') as f:
            data = json.load(f)
        output = sys.argv[2] if len(sys.argv) > 2 else "/tmp/forecast_output.xlsx"
        generate_forecast_excel(data, output)
        print(f"Generated: {output}")
    else:
        print("Usage: python generate_forecast.py <data.json> [output.xlsx]")